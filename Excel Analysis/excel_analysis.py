import os
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from capvalis_first_algorithm import capvalis_first_algorithm

trade_results, daily_metrics, RISK_PER_TRADE, MAX_TRADES_PER_DAY = capvalis_first_algorithm()

# Create Excel file with multiple sheets
try:
    if trade_results:
        # Create output directory if it doesn't exist
        output_dir = r"C:\Users\quick\Desktop\Capvalis\Git"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Generate timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(output_dir, f"Trading_Analysis_{timestamp}.xlsx")
        
        # Create a new workbook
        wb = Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        data_font = Font(name="Calibri")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center')
        
        # 1. Trade Details Sheet
        ws_trades = wb.active
        ws_trades.title = "Trade Details"
        
        # Add headers
        headers = ['Date', 'Time', 'Action', 'Entry Price', 'Stop Loss', 'Target', 
                  'Range High', 'Range Low', 'Position Size', 'Exit Price', 
                  'Exit Time', 'Status', 'P&L']
        
        for col, header in enumerate(headers, 1):
            cell = ws_trades.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment
        
        # Add trade data
        row = 2
        current_month = None
        monthly_pnl = 0

        # Sort trade results by date
        trade_results.sort(key=lambda x: x['date'])

        # Group trades by month
        monthly_trades = {}
        for trade in trade_results:
            date = datetime.strptime(trade['date'], '%Y-%m-%d')
            month_key = date.strftime('%Y-%m')
            if month_key not in monthly_trades:
                monthly_trades[month_key] = []
            monthly_trades[month_key].append(trade)

        # Process each month's trades
        for month_key, trades in monthly_trades.items():
            month_pnl = 0
            
            # Add trades for this month
            for trade in trades:
                ws_trades.cell(row=row, column=1, value=trade['date'])
                ws_trades.cell(row=row, column=2, value=trade['time'])
                ws_trades.cell(row=row, column=3, value=trade['action'])
                ws_trades.cell(row=row, column=4, value=trade['entry'])
                ws_trades.cell(row=row, column=5, value=trade['stop_loss'])
                ws_trades.cell(row=row, column=6, value=trade['target'])
                ws_trades.cell(row=row, column=7, value=trade['range_high'])
                ws_trades.cell(row=row, column=8, value=trade['range_low'])
                ws_trades.cell(row=row, column=9, value=trade['position_size'])
                ws_trades.cell(row=row, column=10, value=trade.get('exit_price', ''))
                ws_trades.cell(row=row, column=11, value=trade.get('exit_time', ''))
                ws_trades.cell(row=row, column=12, value=trade['status'])
                ws_trades.cell(row=row, column=13, value=trade.get('pnl', 0))
                
                # Apply formatting
                for col in range(1, 14):
                    cell = ws_trades.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = alignment
                
                month_pnl += trade.get('pnl', 0)
                row += 1
            
            # Add monthly summary row
            ws_trades.cell(row=row, column=1, value=f"Monthly Total ({month_key})")
            ws_trades.cell(row=row, column=13, value=month_pnl)
            
            # Highlight the monthly summary row
            for col in range(1, 14):
                cell = ws_trades.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = alignment
            
            row += 1

        # 2. Daily Summary Sheet
        ws_daily = wb.create_sheet("Daily Summary")
        
        # Add headers
        headers = ['Date', 'Total Trades', 'Winning Trades', 'Losing Trades', 
                  'Win Rate', 'Daily P&L']
        
        for col, header in enumerate(headers, 1):
            cell = ws_daily.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment
        
        # Add daily metrics and monthly summaries
        current_month = None
        monthly_pnl = 0
        row = 2

        # Sort daily metrics by date
        daily_metrics.sort(key=lambda x: x['date'])

        # Group metrics by month
        monthly_groups = {}
        for metric in daily_metrics:
            date = datetime.strptime(metric['date'], '%Y-%m-%d')
            month_key = date.strftime('%Y-%m')
            if month_key not in monthly_groups:
                monthly_groups[month_key] = []
            monthly_groups[month_key].append(metric)

        # Process each month's data
        for month_key, metrics in monthly_groups.items():
            month_pnl = 0
            
            # Add daily metrics for this month
            for metric in metrics:
                ws_daily.cell(row=row, column=1, value=metric['date'])
                ws_daily.cell(row=row, column=2, value=metric['total_trades'])
                ws_daily.cell(row=row, column=3, value=metric['winning_trades'])
                ws_daily.cell(row=row, column=4, value=metric['losing_trades'])
                ws_daily.cell(row=row, column=5, value=metric['win_rate'])
                ws_daily.cell(row=row, column=6, value=metric['daily_pnl'])
                
                # Apply formatting
                for col in range(1, 7):
                    cell = ws_daily.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = alignment
                
                month_pnl += metric['daily_pnl']
                row += 1
            
            # Add monthly summary row
            ws_daily.cell(row=row, column=1, value=f"Monthly Total ({month_key})")
            ws_daily.cell(row=row, column=6, value=month_pnl)
            
            # Highlight the monthly summary row
            for col in range(1, 7):
                cell = ws_daily.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = alignment
            
            row += 1

        # 3. Performance Metrics Sheet
        ws_metrics = wb.create_sheet("Performance Metrics")
        
        # Calculate overall metrics
        total_pnl = sum(trade.get('pnl', 0) for trade in trade_results)
        total_trades = len(trade_results)
        winning_trades = len([t for t in trade_results if t.get('pnl', 0) > 0])
        losing_trades = len([t for t in trade_results if t.get('pnl', 0) < 0])
        win_rate = winning_trades / total_trades if total_trades > 0 else 0
        
        winning_pnls = [t.get('pnl', 0) for t in trade_results if t.get('pnl', 0) > 0]
        losing_pnls = [t.get('pnl', 0) for t in trade_results if t.get('pnl', 0) < 0]
        
        avg_win = np.mean(winning_pnls) if winning_pnls else 0
        avg_loss = np.mean(losing_pnls) if losing_pnls else 0
        profit_factor = abs(sum(winning_pnls) / sum(losing_pnls)) if losing_pnls else float('inf')
        
        # Add metrics
        metrics = [
            ['Total Trades', total_trades],
            ['Winning Trades', winning_trades],
            ['Losing Trades', losing_trades],
            ['Win Rate', f"{win_rate:.2%}"],
            ['Total P&L', f"₹{total_pnl:,.2f}"],
            ['Average Win', f"₹{avg_win:,.2f}"],
            ['Average Loss', f"₹{avg_loss:,.2f}"],
            ['Profit Factor', f"{profit_factor:.2f}"],
            ['Risk per Trade', f"{RISK_PER_TRADE:.1%}"],
            ['Stop Loss & Target Points', 'Dynamic based on price range'],
            ['Max Trades per Day', MAX_TRADES_PER_DAY]
        ]
        
        for row, (metric, value) in enumerate(metrics, 2):
            ws_metrics.cell(row=row, column=1, value=metric)
            ws_metrics.cell(row=row, column=2, value=value)
            
            # Apply formatting
            for col in range(1, 3):
                cell = ws_metrics.cell(row=row, column=col)
                cell.font = data_font
                cell.border = border
                cell.alignment = alignment
        
        # Apply header formatting to metrics sheet
        for col in range(1, 3):
            cell = ws_metrics.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment
        
        # Adjust column widths
        for ws in [ws_trades, ws_daily, ws_metrics]:
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Save the workbook
        wb.save(excel_file)
        
    else:
        pass
except Exception as e:
    pass

# After the loop completes
print(f"Backtest completed. Results saved to {excel_file}")